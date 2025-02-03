from django.utils import timezone
from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from assistance import models
from utils.dates import get_current_week


# FILTERS


class TodayDateFilter(admin.SimpleListFilter):
    """ Custom detae filter with default value: today """
    title = 'Fecha'
    parameter_name = 'date'

    def lookups(self, request, model_admin):
        return [
            ('today', 'Hoy'),
            ('all', 'Todas las fechas (filtrar por año y semana)'),
        ]

    def queryset(self, request, queryset):
        # Value is the selected filter
        time_zone = timezone.get_current_timezone()
        if self.value() == 'today':
            return queryset.filter(
                date=timezone.now().astimezone(time_zone).date()
            )
        elif self.value() == 'week':
            return queryset.filter(
                date__week=timezone.now().astimezone(
                    time_zone).isocalendar()[1]
            )
        elif self.value() == 'month':
            return queryset.filter(
                date__month=timezone.now().astimezone(time_zone).month
            )
        return queryset

    def value(self):
        # Set default to 'today'
        value = super().value()
        return value or 'today'


class WeekNumberFilter(admin.SimpleListFilter):
    """ Custom filter for week number with default value as the current week """
    title = 'Número de Semana'
    parameter_name = 'week_number'

    def lookups(self, request, model_admin):
        """ Defines the available options in the filter """
        # Get all distinct week numbers from the dataset
        week_numbers = model_admin.get_queryset(
            request).values('week_number').distinct()
        current_week = get_current_week()

        # Generate the options
        options = []
        for week_number in week_numbers:
            if week_number['week_number'] == current_week:
                options.append((
                    week_number['week_number'],
                    f'Semana actual ({current_week})'
                ))
            else:
                options.append((
                    week_number['week_number'],
                    f'Semana {week_number["week_number"]}'
                ))

        return options

    def queryset(self, request, queryset):
        """ Filters the queryset based on the selected value """
        if self.value():
            return queryset.filter(week_number=self.value())
        return queryset

    def value(self):
        """ Sets the default value to the current week number """
        value = super().value()
        if value is None:
            return str(get_current_week())
        return value


class YearFilter(admin.SimpleListFilter):
    title = "Año"
    parameter_name = "year"
    field_name = ""

    def lookups(self, request, model_admin):
        """ Defines the available options in the filter """
        field_name = getattr(model_admin, "year_filter_field", "start_date")
        YearFilter.field_name = field_name
        years = model_admin.get_queryset(request).values(
            f"{YearFilter.field_name}__year"
        ).distinct()
        years_values = [year[f"{YearFilter.field_name}__year"]
                        for year in years]
        years_unique = list(set(years_values))
        options = []
        for year in years_unique:
            options.append((str(year), year))
        return options

    def queryset(self, request, queryset):
        """ Filters the queryset based on the selected value """
        if self.value():
            return queryset.filter(**{f"{YearFilter.field_name}__year": self.value()})
        return queryset

    def value(self):
        """ Sets the default value to the current year """
        value = super().value()
        if value is None:
            return str(timezone.now().year)
        return value


# MODELS


@admin.register(models.Assistance)
class AssistanceAdmin(admin.ModelAdmin):
    """ Assistance model admin """
    list_display = (
        'employee',
        'company',
        'attendance',
        'extra_paid_hours',
        'extra_unpaid_hours',
        'custom_links',
        'notes',
    )
    search_fields = (
        'weekly_assistance__service__agreement__company_name',
        'weekly_assistance__service__location',
        'weekly_assistance__service__description',
        'weekly_assistance__service__employee__name',
        'weekly_assistance__service__employee__last_name_1',
        'weekly_assistance__service__employee__last_name_2',
        'notes',
    )
    list_filter = (
        YearFilter,
        'weekly_assistance__week_number',
        TodayDateFilter,
        'weekly_assistance__service__agreement__company_name',
        'weekly_assistance__service__employee',
        'attendance',
    )
    readonly_fields = (
        'date',
        'weekly_assistance',
    )
    year_filter_field = "date"
    ordering = ('-date',)
    list_editable = ('attendance', 'extra_paid_hours',
                     'extra_unpaid_hours', 'notes')

    # Custom fields
    def company(self, obj):
        """ Return the company name """
        return obj.weekly_assistance.service.agreement.company_name

    def employee(self, obj):
        """ Return the employee name """
        return str(obj.weekly_assistance.service.employee)
    
    def custom_links(self, obj):
        """ Create custom Imprimir and Ver buttons """

        # Generate links
        link_view = "/admin/assistance/extrapayment/?"
        link_view += f"assistance__id__exact={obj.id}"
        
        link_add = "/admin/assistance/extrapayment/add/?"
        link_add += f"assistance={obj.id}"

        return format_html(
            '<a class="btn btn-primary my-1 w-120" href="{}">Ver extras</a>'
            '<a class="btn btn-primary my-1 w-120" href="{}">Añadir extra</a>',
            link_view, link_add
        )

    # Labels for custom fields
    company.short_description = 'Empresa'
    employee.short_description = 'Empleado'
    custom_links.short_description = 'Acciones'


@admin.register(models.WeeklyAssistance)
class WeeklyAssistanceAdmin(admin.ModelAdmin):
    """ Weekly assistance model admin """

    list_display = (
        'company_name',
        'employee',
        'week_number',
        'thursday',
        'friday',
        'saturday',
        'sunday',
        'monday',
        'tuesday',
        'wednesday',
        'total_extra_paid_hours',
        'total_extra_unpaid_hours',
        'custom_links',
    )
    search_fields = (
        'service__agreement__company_name',
        'service__location',
        'service__description',
        'service__employee__name',
        'service__employee__last_name_1',
        'service__employee__last_name_2',
    )
    list_filter = (
        YearFilter,
        WeekNumberFilter,
        'service__agreement__company_name',
        'service__employee',
    )
    readonly_fields = (
        'service',
        'week_number',
        'start_date',
        'end_date',
        'total_extra_paid_hours',
        'total_extra_unpaid_hours',
        'monday',
        'tuesday',
        'wednesday',
        'thursday',
        'friday',
        'saturday',
        'sunday',
        'notes',
    )
    year_filter_field = "start_date"

    # Custom fields
    def company_name(self, obj):
        """ Return the company name """
        return obj.service.agreement.company_name

    def employee(self, obj):
        """ Return the employee name """
        return str(obj.service.employee)

    def custom_links(self, obj):
        """ Create custom Imprimir and Ver buttons """

        employee_id = obj.service.employee.id
        week_number = obj.week_number

        link = "/admin/assistance/assistance/?"
        link += f"weekly_assistance__service__employee__id__exact={
            employee_id}"
        link += f"&weekly_assistance__week_number={week_number}"
        link += f"&year={obj.start_date.year}"
        link += "&date=all"
        return format_html(
            '<a class="btn btn-primary my-1 w-110" href="{}">Editar Dias</a>',
            link
        )

    # Labels for custom fields
    company_name.short_description = 'Empresa'
    employee.short_description = 'Empleado'
    custom_links.short_description = 'Acciones'

    # Custom actions
    def export_excel(self, request, queryset):
        """ Export the queryset to an Excel file with custom styles """

        # Create file and sheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        week_num = queryset.first().week_number
        worksheet.title = f'Asistencias Semana {week_num}'

        # Save sales header
        header = queryset.first().get_data_header()
        worksheet.append(header)

        # Apply header styles (dark grey background, white bold text, larger font size)
        header_fill = PatternFill(
            start_color="404040", end_color="404040", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        # Assuming header is row 1
        for col_num, cell in enumerate(worksheet[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

        # Append data rows and alternate row colors (white and light grey)
        row_fill_white = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        row_fill_grey = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        # Start at row 2 for data
        for row_num, weekly_assistance in enumerate(queryset, start=2):
            data = weekly_assistance.get_data_list()
            worksheet.append(data)

            # Apply alternating row colors
            fill = row_fill_white if row_num % 2 == 0 else row_fill_grey
            for cell in worksheet[row_num]:
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True)

        # Auto-adjust column widths
        for col_num in range(1, len(header) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in worksheet[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 4  # Add some padding
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Set up response
        content_type = "application/vnd.openxmlformats-officedocument"
        content_type += ".spreadsheetml.sheet"
        response = HttpResponse(content_type=content_type)
        response['Content-Disposition'] = 'attachment; filename=export.xlsx'
        workbook.save(response)

        return response

    export_excel.short_description = 'Exportar a Excel'
    actions = [export_excel]


@admin.register(models.ExtraPaymentCategory)
class ExtraPaymentCategoryAdmin(admin.ModelAdmin):
    """ Extra payment category model admin """
    list_display = (
        'name',
        'description',
    )
    

@admin.register(models.ExtraPayment)
class ExtraPaymentAdmin(admin.ModelAdmin):
    """ Extra payment model admin """
    list_display = (
        'assistance',
        'category',
        'amount',
        'notes',
    )
    search_fields = (
        'notes',
    )
    list_filter = (
        'category',
        'assistance',
        'assistance__weekly_assistance__week_number',
        'assistance__weekly_assistance__service__agreement__company_name',
        'assistance__weekly_assistance__service__employee',
    )
    readonly_fields = (
        'created_at',
        'updated_at',
    )