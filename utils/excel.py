from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def get_excel_response(queryset, sheet_name_prefix):
    """ Export a queryset to an Excel file with a single sheet
    
    Args:
        queryset (QuerySet): The queryset to export
        sheet_name (str): The name of the sheet
    """

    # Create file and sheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    week_num = queryset.first().week_number
    worksheet.title = f"{sheet_name_prefix} {week_num}"

    # Save sales header
    header = queryset.first().get_data_header()
    worksheet.append(header)

    # Apply header styles (dark grey background, white bold text, larger font size)
    header_fill = PatternFill(
        start_color="404040", end_color="404040", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=12)
    # Assuming header is row 1
    for col_num, cell in enumerate(worksheet[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Append data rows and alternate row colors (white and light grey)
    row_fill_white = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    row_fill_grey = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )
    row_fill_yellow = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    # Start at row 2 for data
    for row_num, object in enumerate(queryset, start=2):
        data = object.get_data_list()
        is_highlighted = object.get_is_highlighted()
        worksheet.append(data)

        # Apply alternating row colors
        if is_highlighted:
            fill = row_fill_yellow
        else:
            fill = row_fill_white if row_num % 2 == 0 else row_fill_grey
            
        for cell in worksheet[row_num]:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True)

    # Auto-adjust column widths
    for col_num in range(1, len(header) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in worksheet[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 4  # Add some padding
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Set up response
    content_type = "application/vnd.openxmlformats-officedocument"
    content_type += ".spreadsheetml.sheet"
    response = HttpResponse(content_type=content_type)
    response["Content-Disposition"] = "attachment; filename=export.xlsx"
    workbook.save(response)

    return response
